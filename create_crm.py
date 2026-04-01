import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def create_crm_excel(filename="Gayrimenkul_CRM_V3.xlsx"):
    wb = Workbook()
    
    # 1. BİLGİ
    ws_input = wb.active
    ws_input.title = "BİLGİ"
    
    title_font = Font(name='Arial', size=16, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    ws_input.merge_cells('B2:F2')
    cell_title = ws_input['B2']
    cell_title.value = "GAYRİMENKUL CRM SİSTEMİ V3 - GELİŞMİŞ"
    cell_title.font = title_font
    cell_title.fill = header_fill
    cell_title.alignment = Alignment(horizontal="center")
    
    ws_input['B4'] = "Lütfen veri girişi için 'CRM_Giris_Formu.bat' uygulamasını kullanın."
    ws_input['B5'] = "Bu dosya tüm verilerinizin depolandığı ana veritabanıdır."

    # 2. MÜŞTERİ LİSTESİ
    ws_customers = wb.create_sheet("MÜŞTERİ_LİSTESİ")
    customer_headers = [
        "Tarih", "Ad Soyad", "Telefon", "E-posta", "Bütçe", "Talep Türü", 
        "Bölge 1", "Bölge 2", "Bölge 3", "Aciliyet", "Notlar"
    ]
    ws_customers.append(customer_headers)

    # 3. SATILIK KONUT LİSTESİ
    ws_satilik_konut = wb.create_sheet("SATILIK_KONUT")
    konut_headers = [
        "Tarih", "İlan No", "Konut Tipi", "Fiyat", "Bölge/Mahalle", 
        "Oda Sayısı", "Kat", "Manzara", "Havuz", "Bahçe", 
        "Sahibi", "Sahibi Tel", "Durum", "Resim_Klasoru", "Notlar"
    ]
    ws_satilik_konut.append(konut_headers)

    # 4. KİRALIK KONUT LİSTESİ
    ws_kiralik_konut = wb.create_sheet("KİRALIK_KONUT")
    kiralik_headers = [
        "Tarih", "İlan No", "Konut Tipi", "Kira Bedeli", "Bölge/Mahalle", 
        "Oda Sayısı", "Kat", "Manzara", "Havuz", "Bahçe", 
        "Sahibi", "Sahibi Tel", "Durum", "Resim_Klasoru", "Notlar"
    ]
    ws_kiralik_konut.append(kiralik_headers)

    # 5. SATILIK ARSA LİSTESİ
    ws_satilik_arsa = wb.create_sheet("SATILIK_ARSA")
    arsa_headers = [
        "Tarih", "İlan No", "Arsa Tipi", "Ada", "Parsel", "Fiyat", "Bölge/Mahalle", 
        "Sahibi", "Sahibi Tel", "Durum", "Resim_Klasoru", "Notlar"
    ]
    ws_satilik_arsa.append(arsa_headers)
    
    # 6. HATIRLATICILAR
    ws_reminders = wb.create_sheet("HATIRLATICILAR")
    reminder_headers = [
        "Tarih", "Müşteri Adı", "Telefon", "Hatırlatma Tarihi", "Not", "Durum"
    ]
    ws_reminders.append(reminder_headers)

    # Styling all headers
    for sheet in [ws_customers, ws_satilik_konut, ws_kiralik_konut, ws_satilik_arsa, ws_reminders]:
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        for col in sheet.columns:
            sheet.column_dimensions[col[0].column_letter].width = 20

    wb.save(filename)
    print(f"'{filename}' başarıyla oluşturuldu.")

if __name__ == "__main__":
    create_crm_excel()
